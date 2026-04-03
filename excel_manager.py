import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "cash_register/daily_info.xlsx"

HEADERS = [
    "Date and Time",     
    "Cashier",            
    "HUMO",              
    "UZCARD",           
    "NAQD",               
    "Rasxod Total",      
    "Expenses Details",   
    "Description",       
    "Additional Info",    
    "1C_Quantity_Sold",   
    "1C_Total_Sales"      
]

def get_file_path() -> str:
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, EXCEL_FILE)

def ensure_excel_exists() -> None:
    path = get_file_path()
    
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reports"
        ws.append(HEADERS)
        wb.save(path)

def append_to_excel(submission: dict) -> None:
    ensure_excel_exists()
    
    path = get_file_path()
    
    wb = load_workbook(path)
    ws = wb.active 
    
    expense_details = []
    for item in submission.get("rasxod_items", []):
        expense_details.append(f"{item['reason']}:{item['amount']}")
    
    expenses_str = "|".join(expense_details)
    
    row = [
        submission.get("submitted_at_local", ""), 
        submission.get("cashier", ""),            
        submission.get("humo", 0.0),              
        submission.get("uzcard", 0.0),            
        submission.get("naqd", 0.0),              
        submission.get("rasxod_total", 0.0),      
        expenses_str,                             
        submission.get("description", ""),        
        "",                                       
        submission.get("sum_quantity", 0),        
        submission.get("sum_sales", 0)            
    ]
    
    ws.append(row)
    
    wb.save(path)
